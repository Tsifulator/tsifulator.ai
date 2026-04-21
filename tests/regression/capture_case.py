#!/usr/bin/env python3
"""capture_case.py — turn a live tsifl run into a new regression case.

Usage (flags):
    python capture_case.py --name 002-comps-table \
        --start /path/to/start.xlsx \
        --expected /path/to/expected.xlsx \
        --message "complete these steps" \
        --image /path/to/screen1.png --image /path/to/screen2.png

Usage (interactive — prompts for anything missing):
    python capture_case.py

What it does:
  1. Creates tests/regression/cases/<name>/
  2. Copies start.xlsx + expected.xlsx in for reference/inspection
  3. Copies any --image files into images/
  4. Reads start.xlsx to build a realistic `context` payload
  5. POSTs the request to /chat, captures the response
  6. Writes request.json (with images) + response.json
  7. Auto-generates a starter rubric.yaml locking in what tsifl emitted
  8. Runs the new case through run_tests.py to verify the rubric passes
  9. Prints a summary + "tune the rubric" next-step hints

You still need to verify the output is actually correct before committing the
case. The rubric locks in behavior — garbage in, garbage locked-in.
"""
from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

try:
    import requests
    import yaml
    from openpyxl import load_workbook
except ImportError as e:
    print(f"Missing dependency: {e}. Run: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(2)

HERE = Path(__file__).resolve().parent
CASES_DIR = HERE / "cases"
DEFAULT_BACKEND = "https://focused-solace-production-6839.up.railway.app"

# Sheet names the LLM has hallucinated in the past — add to forbidden list by default
COMMON_PHANTOM_SHEETS = [
    "Transactions", "Transactions Stats", "Summary", "Criteria",
    "Data", "Stats", "Sheet1",
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def c_green(s): return f"\033[92m{s}\033[0m"
def c_red(s):   return f"\033[91m{s}\033[0m"
def c_dim(s):   return f"\033[90m{s}\033[0m"
def c_bold(s):  return f"\033[1m{s}\033[0m"


def ask(prompt: str, default: str | None = None, required: bool = True) -> str:
    hint = f" [{default}]" if default else ""
    while True:
        raw = input(f"{prompt}{hint}: ").strip()
        if raw:
            return raw
        if default is not None:
            return default
        if not required:
            return ""
        print("  (required)")


def slug(s: str) -> str:
    return re.sub(r"[^a-z0-9-]+", "-", s.lower()).strip("-")


def _cell_to_json(v: Any) -> Any:
    """Coerce a cell value to something json.dumps can handle.

    openpyxl returns ArrayFormula / datetime / decimal objects for some cells.
    The Excel add-in would serialize these as display strings, so we do the same.
    """
    import datetime, decimal
    if v is None or isinstance(v, (bool, int, float, str)):
        return v
    # openpyxl array formula
    if hasattr(v, "text"):
        return getattr(v, "text", str(v))
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    if isinstance(v, decimal.Decimal):
        return float(v)
    return str(v)


def build_context_from_workbook(path: Path, active: str | None = None) -> dict:
    """Build a `context` dict mimicking the Excel add-in's getExcelContext()."""
    wb = load_workbook(path, data_only=False)
    active = active or wb.worksheets[0].title
    sheet_summaries = []
    active_data: list[list[Any]] = []
    for ws in wb.worksheets:
        preview_rows = 200 if ws.title == active else 100
        preview = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(preview_rows, ws.max_row or 1),
            max_col=min(26, ws.max_column or 1),
        ):
            preview.append([_cell_to_json(c.value) for c in row])
        sheet_summaries.append({
            "name": ws.title,
            "used_range": ws.calculate_dimension() if ws.max_row else "empty",
            "rows": ws.max_row or 0,
            "cols": ws.max_column or 0,
            "preview": preview,
        })
        if ws.title == active:
            active_data = preview

    active_ws = wb[active]
    # Match the Excel add-in's namedRanges shape: list of {name, reference}
    named_ranges = []
    for nm in wb.defined_names:
        try:
            named_ranges.append({"name": nm, "reference": wb.defined_names[nm].value})
        except Exception:
            named_ranges.append({"name": nm, "reference": ""})

    return {
        "app":              "excel",
        "sheet":            active,
        "all_sheets":       [ws.title for ws in wb.worksheets],
        "selected_cell":    "A1",
        "selected_value":   None,
        "used_range":       active_ws.calculate_dimension() if active_ws.max_row else "empty",
        "sheet_data":       active_data,
        "sheet_formulas":   active_data,
        "sheet_summaries":  sheet_summaries,
        "named_ranges":     named_ranges,
        "preferences":      {},
    }


def encode_images(paths: list[Path]) -> list[dict]:
    out = []
    for p in paths:
        if not p.exists():
            print(c_red(f"  image not found: {p}"), file=sys.stderr)
            continue
        mime, _ = mimetypes.guess_type(str(p))
        out.append({
            "file_name": p.name,
            "media_type": mime or "image/png",
            "data": base64.b64encode(p.read_bytes()).decode("ascii"),
        })
    return out


# ── Rubric generation ────────────────────────────────────────────────────────

def _action_targets(actions: list[dict]) -> list[str]:
    out = []
    for a in actions:
        t = a.get("type", "")
        p = a.get("payload") or {}
        s = p.get("sheet")
        if t in ("write_cell", "write_formula"):
            cell = p.get("cell") or p.get("address")
            if cell:
                out.append(f"{s}!{cell}" if s else cell)
        elif t in ("write_range", "format_range", "clear_range", "fill_down", "fill_right"):
            rng = p.get("range")
            if rng:
                out.append(f"{s}!{rng}" if s else rng)
        elif t == "create_named_range":
            ref = p.get("reference") or ""
            if ref:
                out.append(ref)
    return out


def _key_formula_cells(actions: list[dict], max_cells: int = 3) -> dict[str, list[str]]:
    """Pick up to N formula-emitting actions and extract tokens from their formulas
    so the rubric ensures the formula structure sticks. Returns a dict keyed by
    'Sheet!Cell' with a list of must-contain substrings."""
    out: dict[str, list[str]] = {}
    for a in actions:
        if a.get("type") != "write_formula":
            continue
        p = a.get("payload") or {}
        sheet = p.get("sheet")
        cell = p.get("cell") or p.get("address")
        formula = str(p.get("formula") or "").strip()
        if not (sheet and cell and formula and formula.startswith("=")):
            continue
        # Extract cell references (A1, $A$1, A1:B10) as anchor tokens
        refs = re.findall(r"\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?", formula)
        if not refs:
            continue
        addr = f"{sheet}!{cell}"
        # Take the first 2 distinct refs as "must contain"
        uniq = list(dict.fromkeys(refs))[:2]
        out[addr] = uniq
        if len(out) >= max_cells:
            break
    return out


def build_rubric(
    actions: list[dict],
    context: dict,
    project_name: str,
    reply: str,
) -> dict:
    action_types = sorted({a.get("type") for a in actions if a.get("type")})
    targets      = _action_targets(actions)
    all_sheets   = context.get("all_sheets") or []

    rubric: dict[str, Any] = {
        "_project": project_name,
        "allowed_sheet_targets":   list(all_sheets),
        "forbidden_sheet_targets": [s for s in COMMON_PHANTOM_SHEETS if s not in all_sheets],
        "min_action_count":        max(1, int(len(actions) * 0.7)),  # 30% slack
        "reply_not_empty":         bool(reply.strip()),
        "must_include_action_types": action_types,
        "must_reference_cells":      sorted(set(targets)),
    }
    formula_checks = _key_formula_cells(actions)
    if formula_checks:
        rubric["formula_at_cell_contains"] = formula_checks
    return rubric


def dump_rubric_yaml(rubric: dict, path: Path, project_name: str) -> None:
    """Write rubric.yaml with comments so the user knows what to tune."""
    header = f"""# {project_name} — auto-generated by capture_case.py
#
# This rubric was captured from a live run. Tune it before committing:
#
#   1. Review `must_reference_cells` — drop entries that are accidental
#      (e.g. the LLM wrote a cell you don't care about) or expand ranges
#      if you want the check broader.
#   2. Review `must_include_action_types` — drop noise types (navigate_sheet,
#      set_number_format) if you only care about structural actions.
#   3. Tighten `formula_at_cell_contains` — for business-critical formulas,
#      add tokens that MUST appear (e.g. specific function names, cell refs).
#   4. Add `formula_at_cell_forbids` for known bad patterns (guessed rates,
#      off-by-one shortcuts, etc.) — these catch regressions most cleanly.
#   5. Add `must_not_overwrite` for cell ranges the LLM should never touch
#      (existing labels, data rows, the user's own work).
#
# After tuning, run `python run_tests.py --case {Path(path).parent.name}`
# and make sure it still passes. If it doesn't, fix the rubric or the backend.

"""
    body = yaml.safe_dump(rubric, sort_keys=False, default_flow_style=False, allow_unicode=True)
    path.write_text(header + body)


# ── Main capture flow ────────────────────────────────────────────────────────

def capture(args) -> int:
    name     = args.name or ask("Case name (e.g. 002-comps-table)")
    start    = Path(args.start    or ask("Start workbook path"))
    expected = Path(args.expected or ask("Expected workbook path (can match start if no target yet)"))
    message  = args.message       or ask("User message (or path to .txt)")
    if Path(message).is_file():
        message = Path(message).read_text()

    images = [Path(p) for p in (args.image or [])]
    if not args.image and not args.no_images_prompt:
        img_input = ask("Image paths (comma-separated, blank for none)", default="", required=False)
        if img_input.strip():
            images = [Path(p.strip()) for p in img_input.split(",") if p.strip()]

    slug_name = slug(name)
    if not slug_name[0].isdigit():
        existing = sorted(d.name for d in CASES_DIR.iterdir() if d.is_dir()) if CASES_DIR.exists() else []
        next_num = len(existing) + 1
        slug_name = f"{next_num:03d}-{slug_name}"
        print(c_dim(f"  normalized case name → {slug_name}"))

    case_dir = CASES_DIR / slug_name
    if case_dir.exists():
        print(c_red(f"Case already exists: {case_dir}"))
        return 2

    case_dir.mkdir(parents=True)
    (case_dir / "images").mkdir()

    # Copy reference workbooks
    shutil.copy(start,    case_dir / "start.xlsx")
    shutil.copy(expected, case_dir / "expected.xlsx")

    # Copy images
    for img in images:
        shutil.copy(img, case_dir / "images" / img.name)

    # Build request
    print(c_dim(f"  building context from {start}…"))
    context = build_context_from_workbook(start)
    request = {
        "user_id": args.user or "regression-test-user",
        "message": message,
        "context": context,
    }
    if images:
        request["images"] = encode_images([case_dir / "images" / i.name for i in images])

    # Hit /chat
    print(c_dim(f"  POST {args.backend}/chat/ …"))
    try:
        resp = requests.post(f"{args.backend.rstrip('/')}/chat/", json=request, timeout=args.timeout)
    except requests.RequestException as e:
        print(c_red(f"request failed: {e}"))
        shutil.rmtree(case_dir)
        return 2

    if resp.status_code != 200:
        print(c_red(f"HTTP {resp.status_code}: {resp.text[:300]}"))
        shutil.rmtree(case_dir)
        return 2

    data    = resp.json()
    actions = data.get("actions") or []
    reply   = data.get("reply") or ""
    print(c_dim(f"  got {len(actions)} action(s), reply: {reply[:80]!r}"))

    if len(actions) == 0 and not data.get("cu_session_id"):
        print(c_red("no actions returned — can't build rubric. Aborting."))
        shutil.rmtree(case_dir)
        return 2

    # Strip base64 image bytes from request.json saved on disk — keep file small.
    # The images live in images/ and are re-attached at test-run time.
    slim_request = dict(request)
    slim_request.pop("images", None)
    (case_dir / "request.json").write_text(json.dumps(slim_request, indent=2, default=str))
    (case_dir / "response.json").write_text(json.dumps(data, indent=2))

    # Auto-generate rubric
    rubric = build_rubric(actions, context, project_name=name, reply=reply)
    dump_rubric_yaml(rubric, case_dir / "rubric.yaml", project_name=name)

    # Sanity-check: run the new case through run_tests.py
    print(c_dim(f"  verifying rubric against the captured response…"))
    verify = subprocess.run(
        [sys.executable, str(HERE / "run_tests.py"),
         "--case", slug_name, "--backend", args.backend],
        capture_output=True, text=True,
    )
    if verify.returncode != 0:
        print(c_red("  verification FAILED — rubric needs manual tuning."))
        print(verify.stdout)
        print(c_dim(f"  case kept at {case_dir} for you to fix."))
        return 1

    print()
    print(c_green(f"✓ Captured case: {slug_name}"))
    print(f"  {case_dir}")
    print()
    print(c_bold("Next steps:"))
    print(f"  1. Open {case_dir / 'rubric.yaml'} and tune — add forbidden patterns,")
    print(f"     must_not_overwrite ranges, or tighten formula checks.")
    print(f"  2. Run: python run_tests.py --case {slug_name}")
    print(f"  3. Once you're confident, commit and push. CI will enforce it.")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--name")
    ap.add_argument("--start", help="start workbook path (.xlsx)")
    ap.add_argument("--expected", help="expected/target workbook path (.xlsx)")
    ap.add_argument("--message", help="user message, or path to a .txt file")
    ap.add_argument("--image", action="append", help="image path (repeatable)")
    ap.add_argument("--user",  default="regression-test-user")
    ap.add_argument("--backend", default=os.environ.get("TSIFL_BACKEND") or DEFAULT_BACKEND)
    ap.add_argument("--timeout", type=int, default=180)
    ap.add_argument("--no-images-prompt", action="store_true",
                    help="skip the interactive images prompt (for scripted use)")
    return capture(ap.parse_args())


if __name__ == "__main__":
    sys.exit(main())
